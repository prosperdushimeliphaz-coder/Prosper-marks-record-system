import streamlit as st
import pandas as pd
from datetime import datetime
import os
from io import BytesIO
from openpyxl import load_workbook, Workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

# =========================
# 🔹 APP CONFIG
# =========================
st.set_page_config(page_title="Student Marks Recorder", layout="centered")

st.title("📘 Student Marks Recorder")
st.write("Enter test scores, save progress, and later generate full reports (Excel or PDF).")

# =========================
# 🔹 SECTION 1: HEADER INFO
# =========================
st.subheader("🏫 School Information")
col1, col2 = st.columns(2)
district = col1.text_input("District")
school = col2.text_input("School Name")
col3, col4 = st.columns(2)
class_name = col3.text_input("Class Name (e.g., S1B)")
academic_year = col4.text_input("Academic Year (e.g., 2025-2026)")

# =========================
# 🔹 SECTION 2: FILE UPLOAD
# =========================
st.subheader("📁 Upload Class Excel Workbook")
uploaded_file = st.file_uploader("Upload Excel file (.xlsx)", type=["xlsx"])

# Temporary storage directory
SAVE_DIR = "saved_data"
os.makedirs(SAVE_DIR, exist_ok=True)
SAVE_PATH = os.path.join(SAVE_DIR, f"{class_name}_marks.xlsx") if class_name else None

# =========================
# 🔹 SECTION 3: ENTER TEST DETAILS
# =========================
if uploaded_file:
    try:
        # Read workbook and list sheets
        excel_data = pd.ExcelFile(uploaded_file)
        sheet_name = st.selectbox("Select class sheet", excel_data.sheet_names)
        df = pd.read_excel(excel_data, sheet_name=sheet_name)

        if 'Name' not in df.columns:
            st.error("❌ 'Name' column not found in your Excel file. Please ensure it has a 'Name' header.")
        else:
            st.success(f"✅ Loaded {len(df)} students from sheet '{sheet_name}'")

            # Select test info
            test_number = st.selectbox("Select Test Number", ["Test 1", "Test 2", "Test 3", "Test 4", "Test 5"])
            test_date = st.date_input("Select test date")
            test_max = st.number_input("Enter maximum marks for this test", min_value=1, max_value=100, value=20)

            st.subheader(f"📝 Enter Scores for {test_number}")
            scores = {}

            for i, row in df.iterrows():
                name = row["Name"]
                score = st.number_input(f"{name}", min_value=0.0, max_value=float(test_max), step=1.0, key=f"{test_number}_{i}")
                scores[name] = score

            # =========================
            # 🔹 SAVE OR UPDATE WORKBOOK
            # =========================
            if st.button("💾 Save/Update This Test"):
                if os.path.exists(SAVE_PATH):
                    workbook = load_workbook(SAVE_PATH)
                else:
                    workbook = Workbook()
                    workbook.remove(workbook.active)
                    workbook.create_sheet(sheet_name)

                ws = workbook[sheet_name]

                # If empty, add headers
                if ws.max_row == 1 and ws.cell(row=1, column=1).value is None:
                    headers = ["S/N", "Name"]
                    ws.append(headers)
                    for i, name in enumerate(df["Name"], start=1):
                        ws.append([i, name])

                # Find column index for this test
                test_col = None
                for cell in ws[1]:
                    if cell.value == test_number:
                        test_col = cell.column
                        break
                if not test_col:
                    ws.cell(row=1, column=ws.max_column + 1, value=test_number)
                    ws.cell(row=2, column=ws.max_column, value=f"({test_date})")
                    ws.cell(row=3, column=ws.max_column, value=f"Max: {test_max}")
                    test_col = ws.max_column

                # Fill scores
                for row_idx in range(4, ws.max_row + 1):
                    name = ws.cell(row=row_idx, column=2).value
                    if name in scores:
                        ws.cell(row=row_idx, column=test_col, value=scores[name])

                workbook.save(SAVE_PATH)
                st.success(f"✅ {test_number} saved successfully to {SAVE_PATH}")

            # =========================
            # 🔹 GENERATE FINAL REPORT
            # =========================
            if os.path.exists(SAVE_PATH):
                if st.button("📊 Generate Full Report"):
                    wb = load_workbook(SAVE_PATH)
                    ws = wb[sheet_name]

                    # Find test columns
                    tests = [cell.value for cell in ws[1] if "Test" in str(cell.value)]
                    max_marks = []
                    for c in range(3, ws.max_column + 1):
                        val = ws.cell(row=3, column=c).value
                        if val and "Max" in str(val):
                            max_marks.append(int(val.split(":")[1].strip()))

                    # Calculate totals and % 
                    ws.cell(row=1, column=ws.max_column + 1, value="Total")
                    ws.cell(row=1, column=ws.max_column + 2, value="%")

                    for r in range(4, ws.max_row + 1):
                        marks = [ws.cell(row=r, column=c).value for c in range(3, ws.max_column - 2) if isinstance(ws.cell(row=r, column=c).value, (int, float))]
                        total = sum(marks)
                        ws.cell(row=r, column=ws.max_column - 1, value=total)
                        percent = round((total / sum(max_marks)) * 100, 2) if sum(max_marks) else 0
                        ws.cell(row=r, column=ws.max_column, value=percent)

                    report_bytes = BytesIO()
                    wb.save(report_bytes)
                    report_bytes.seek(0)

                    st.download_button("⬇️ Download Excel Report", data=report_bytes, file_name=f"{class_name}_Final_Report.xlsx", mime="application/vnd.ms-excel")

                    # ---- Generate PDF
                    pdf_bytes = BytesIO()
                    c = canvas.Canvas(pdf_bytes, pagesize=A4)
                    c.setFont("Helvetica-Bold", 14)
                    c.drawString(200, 800, "STUDENT MARKS REPORT")
                    c.setFont("Helvetica", 11)
                    c.drawString(50, 780, f"District: {district}")
                    c.drawString(50, 765, f"School: {school}")
                    c.drawString(50, 750, f"Class: {class_name}")
                    c.drawString(50, 735, f"Academic Year: {academic_year}")
                    c.drawString(50, 715, f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}")

                    c.setFont("Helvetica-Bold", 12)
                    y = 690
                    c.drawString(50, y, "Name")
                    c.drawString(300, y, "Total")
                    c.drawString(400, y, "%")
                    c.setFont("Helvetica", 11)
                    y -= 20

                    for r in range(4, ws.max_row + 1):
                        name = str(ws.cell(row=r, column=2).value)
                        total = ws.cell(row=r, column=ws.max_column - 1).value
                        percent = ws.cell(row=r, column=ws.max_column).value
                        c.drawString(50, y, name[:30])
                        c.drawString(300, y, str(total))
                        c.drawString(400, y, f"{percent}%")
                        y -= 15
                        if y < 100:
                            c.showPage()
                            y = 780

                    c.save()
                    pdf_bytes.seek(0)
                    st.download_button("⬇️ Download PDF Report", data=pdf_bytes, file_name=f"{class_name}_Final_Report.pdf", mime="application/pdf")

    except Exception as e:
        st.error(f"⚠️ Error: {e}")
